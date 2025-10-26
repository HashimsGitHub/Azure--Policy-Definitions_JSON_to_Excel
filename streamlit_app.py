import streamlit as st
import io
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Page Configuration ---
st.set_page_config(page_title="Azure Policy Definitions - JSON to Excel", page_icon="🎈", layout="wide")

# --- Header Section ---
st.markdown("""
    <div style="display:flex; align-items:center; gap:15px; background-color:#0078D4; padding:15px; border-radius:10px; color:white;">
        <img src="https://upload.wikimedia.org/wikipedia/commons/f/fa/Microsoft_Azure.svg" width="60">
        <h1 style="margin:0;">Azure Policy Definitions - JSON to Excel Converter</h1>
    </div>
    <p style="margin-top:10px; font-size:16px; color:#333;">
        Upload your Azure Policy Definitions (JSON export) and generate a clean, formatted Excel report 📊
    </p>
""", unsafe_allow_html=True)

# --- Region Formatter ---
def format_location(loc: str) -> str:
    if not loc:
        return ""
    loc_lower = loc.lower()
    region_map = {
        "australiaeast": "Australia East", "australiasoutheast": "Australia Southeast",
        "northeurope": "North Europe", "westeurope": "West Europe",
        "eastus": "East US", "eastus2": "East US 2",
        # Add more as needed
    }
    return region_map.get(loc_lower, loc)

# --- Extract the last part of the Policy ID ---
def extract_policy_id(full_id: str) -> str:
    """Extract the GUID part from the full policy ID."""
    return full_id.split('/')[-1] if full_id else ""

# --- File Upload Section ---
uploaded_file = st.file_uploader("📂 Upload your Azure Policy Definitions JSON file", type="json")

if uploaded_file is not None:
    # --- Parse JSON ---
    data = json.load(uploaded_file)
    records = []
    
    for policy in data:
        description = policy.get("description", "")
        display_name = policy.get("displayName", "")
        policy_id = extract_policy_id(policy.get("id", ""))  # Updated here
        category = policy.get("metadata", {}).get("category", "")
        policy_type = policy.get("policyType", "")
        effect = policy.get("policyRule", {}).get("then", {}).get("effect", "None")
        versions = ", ".join(policy.get("versions", []))
        metadata = {
            "Policy ID": policy_id,  # Now showing only the GUID
            "Display Name": display_name,
            "Description": description,
            "Category": category,
            "Policy Type": policy_type,
            "Effect": effect,
            "Versions": versions
        }
        records.append(metadata)    

    # --- Create DataFrame ---
    df_policies = pd.DataFrame(records)

    # Add an index column (starting from 1)
    df_policies.insert(0, 'Index', range(1, len(df_policies) + 1))
    
    # --- Excel File Creation ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Policy_Definitions"
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    title_font = Font(bold=True, size=14)
    align_center = Alignment(horizontal="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    
    # --- Add Title and Data to Excel ---
    ws.append(["Azure Policy Definitions"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A1"].font = title_font
    ws["A1"].alignment = align_center
    ws.append([""])

    # Add header row with 'Index'
    ws.append(["Index", "Policy ID", "Display Name", "Description", "Category", "Policy Type", "Effect", "Versions"])

    # Style headers
    for i in range(1, len(df_policies.columns) + 1):
        cell = ws[f"{get_column_letter(i)}{ws.max_row}"]
        cell.font = bold
        cell.fill = hdr_fill
    
    # Add the rows to Excel
    for row in df_policies.itertuples(index=False):
        ws.append(list(row))

    # --- Apply borders and column width adjustments ---
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.border = border
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(len(str(c.value)) for c in col if c.value) + 3

    # Save Excel file to memory
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # --- Display Tables on Streamlit UI ---
    st.markdown("### 📘 Policy Definitions Overview")
    st.dataframe(df_policies, use_container_width=True)

    # --- Download Button ---
    st.download_button(
        label="💾 Download Excel Report",
        data=excel_file,
        file_name="Azure_Policy_Definitions.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Click to download the formatted Excel report."
    )

else:
    st.info("👆 Please upload a JSON file to begin.")

# --- Footer ---
st.markdown("""
    <hr style="margin-top:40px;">
    <div style="text-align:center; color:gray; font-size:14px;">
        Built with ❤️ using Streamlit & Microsoft Azure<br>
        © 2025 Hashim Hilal — Cloud Architect 
    </div>
""", unsafe_allow_html=True)
